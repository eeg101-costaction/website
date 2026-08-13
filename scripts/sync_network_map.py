#!/usr/bin/env python3
"""Validate an eCOST member export and rebuild the EEG101 Network Map dataset.

The script is deliberately conservative. It refuses to overwrite the public map
when the export is incomplete, unexpectedly small, or contains institutions that
have no reviewed location record. This protects the public map from bad exports
and avoids placing new institutions at an inaccurate location.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import re
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path
from typing import Any, Iterable


REQUIRED_FIELDS = {"first name", "last name", "email", "affiliation", "country"}
GROUP_FIELD_CANDIDATES = {"assigned working groups", "application working groups"}
GROUP_PATTERN = re.compile(r"\bWG\s*([123])\b", re.IGNORECASE)
COUNTRY_SUFFIX_PATTERN = re.compile(r"\s*\([A-Z]{2,3}\)\s*$")
COUNTRY_ALIASES = {"Turkey": "Türkiye", "Turkiye": "Türkiye"}


def clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def field_key(value: str) -> str:
    return re.sub(r"\s+", " ", clean(value).lstrip("\ufeff").lower())


def ascii_key(value: str) -> str:
    text = unicodedata.normalize("NFKD", clean(value)).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", " ", text.lower()).strip()


def normalise_country(value: str) -> str:
    country = COUNTRY_SUFFIX_PATTERN.sub("", clean(value))
    return COUNTRY_ALIASES.get(country, country)


def obfuscate_email(value: str) -> str:
    email = clean(value)
    if "@" not in email:
        return ""
    local, domain = email.split("@", 1)
    return f"{local} [at] {domain.replace('.', ' [dot] ')}"


def coerce_rows_from_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        preview = handle.read(4096)
        handle.seek(0)
        try:
            dialect = csv.Sniffer().sniff(preview, delimiters=";,\t")
        except csv.Error:
            dialect = csv.excel
            dialect.delimiter = ";"
        return [
            {field_key(key): clean(value) for key, value in row.items() if key is not None}
            for row in csv.DictReader(handle, dialect=dialect)
        ]


def coerce_rows_from_xlsx(path: Path) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("The .xlsx export requires openpyxl.") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    headers = [field_key(value) for value in next(rows)]
    return [
        {headers[index]: clean(value) for index, value in enumerate(row) if index < len(headers)}
        for row in rows
        if any(clean(value) for value in row)
    ]


def load_rows(path: Path) -> list[dict[str, str]]:
    if path.suffix.lower() == ".csv":
        return coerce_rows_from_csv(path)
    if path.suffix.lower() == ".xlsx":
        return coerce_rows_from_xlsx(path)
    raise RuntimeError("eCOST export must be a .csv or .xlsx file.")


def extract_groups(row: dict[str, str]) -> list[str]:
    text = " ".join(row.get(field, "") for field in GROUP_FIELD_CANDIDATES)
    groups = {f"WG{group}" for group in GROUP_PATTERN.findall(text)}
    for number in (1, 2, 3):
        value = row.get(f"wg{number}. reporting standards", "") if number == 1 else ""
        if number == 2:
            value = row.get("wg2. curation and harmonization", "")
        if number == 3:
            value = row.get("wg3. manifesto", "")
        if clean(value).lower() in {"y", "yes", "true", "1"}:
            groups.add(f"WG{number}")
    return sorted(groups, key=lambda group: int(group[-1]))


def build_location_index(current: dict[str, Any]) -> dict[tuple[str, str], dict[str, Any]]:
    locations: dict[tuple[str, str], dict[str, Any]] = {}
    for site in current.get("sites", []):
        country = normalise_country(site["country"])
        location = {
            "institution": site["institution"],
            "city": site["city"],
            "country": country,
            "latitude": site["latitude"],
            "longitude": site["longitude"],
            "location_confidence": site.get("location_confidence", "reviewed"),
        }
        aliases = {site.get("institution", "")}
        aliases.update(member.get("affiliation", "") for member in site.get("members", []))
        for alias in aliases:
            key = (ascii_key(alias), ascii_key(country))
            if key[0]:
                locations[key] = location
    return locations


def validate_headers(rows: list[dict[str, str]]) -> None:
    if not rows:
        raise RuntimeError("The eCOST export contains no member rows.")
    headers = set(rows[0])
    missing = sorted(REQUIRED_FIELDS - headers)
    has_group_source = bool(headers & GROUP_FIELD_CANDIDATES) or any(
        field.startswith("wg1.") or field.startswith("wg2.") or field.startswith("wg3.")
        for field in headers
    )
    if missing or not has_group_source:
        details = []
        if missing:
            details.append("missing fields: " + ", ".join(missing))
        if not has_group_source:
            details.append("no Working Group assignment fields")
        raise RuntimeError(
            "The downloaded export cannot safely rebuild the public map (" + "; ".join(details) + "). "
            "Use the detailed eCOST member export that includes affiliations, countries, and Working Group assignments."
        )


def map_members(rows: Iterable[dict[str, str]], locations: dict[tuple[str, str], dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, str]]]:
    grouped: dict[tuple[str, str], dict[str, Any]] = {}
    unresolved: list[dict[str, str]] = []

    for row in rows:
        first = clean(row.get("first name"))
        last = clean(row.get("last name"))
        name = " ".join(part for part in (first, last) if part)
        email = clean(row.get("email"))
        affiliation = clean(row.get("affiliation"))
        country = normalise_country(row.get("country", ""))
        if not name or not email or not affiliation or not country:
            raise RuntimeError("A required member value is blank in the eCOST export.")

        location_key = (ascii_key(affiliation), ascii_key(country))
        location = locations.get(location_key)
        if location is None:
            unresolved.append({"name": name, "affiliation": affiliation, "country": country})
            continue

        member = {
            "name": name,
            "affiliation": affiliation,
            "working_groups": extract_groups(row),
            "email": obfuscate_email(email),
            "homepage": clean(row.get("homepages")),
            "orcid": clean(row.get("orcid")),
        }
        site_key = (ascii_key(location["institution"]), ascii_key(location["country"]))
        if site_key not in grouped:
            grouped[site_key] = {"location": location, "members": []}
        grouped[site_key]["members"].append(member)

    sites: list[dict[str, Any]] = []
    ordered_groups = sorted(
        grouped.values(),
        key=lambda group: (group["location"]["institution"].casefold(), group["location"]["country"].casefold()),
    )
    for index, group in enumerate(ordered_groups, start=1):
        location = group["location"]
        members = group["members"]
        members.sort(key=lambda member: member["name"].casefold())
        working_groups = sorted({tag for member in members for tag in member["working_groups"]}, key=lambda tag: int(tag[-1]))
        sites.append({
            "id": f"site-{index}",
            **location,
            "members": members,
            "member_count": len(members),
            "working_groups": working_groups,
        })
    return sites, unresolved


def build_dataset(export_path: Path, existing_path: Path, output_path: Path, report_path: Path) -> None:
    rows = load_rows(export_path)
    validate_headers(rows)
    current = json.loads(existing_path.read_text(encoding="utf-8"))
    baseline_count = int(current.get("member_count", 0))
    minimum_count = max(20, int(baseline_count * 0.7))
    if len(rows) < minimum_count:
        raise RuntimeError(
            f"Export contains {len(rows)} members, below the safety threshold of {minimum_count}."
        )

    locations = build_location_index(current)
    sites, unresolved = map_members(rows, locations)
    report = {
        "export": export_path.name,
        "export_member_count": len(rows),
        "baseline_member_count": baseline_count,
        "resolved_member_count": sum(site["member_count"] for site in sites),
        "unresolved_institutions": unresolved,
        "generated_at": dt.datetime.now(dt.timezone.utc).replace(microsecond=0).isoformat(),
    }
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(json.dumps(report, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    if unresolved:
        raise RuntimeError(
            f"{len(unresolved)} members have institutions without reviewed map locations. "
            f"See {report_path}; the public map was not changed."
        )

    countries: dict[str, dict[str, Any]] = {}
    for site in sites:
        country = countries.setdefault(site["country"], {"name": site["country"], "site_count": 0, "member_count": 0})
        country["site_count"] += 1
        country["member_count"] += site["member_count"]

    dataset = {
        "generated_from": export_path.name,
        "member_count": len(rows),
        "site_count": len(sites),
        "country_count": len(countries),
        "countries": sorted(countries.values(), key=lambda country: country["name"]),
        "sites": sites,
    }
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(dataset, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(json.dumps({"members": len(rows), "sites": len(sites), "countries": len(countries), "report": str(report_path)}, ensure_ascii=False))


def main() -> int:
    parser = argparse.ArgumentParser(description="Synchronise a detailed eCOST export to the EEG101 Network Map")
    parser.add_argument("--input", type=Path, required=True, help="Detailed eCOST .csv or .xlsx export")
    parser.add_argument("--existing", type=Path, default=Path("assets/data/network-map.json"))
    parser.add_argument("--output", type=Path, default=Path("assets/data/network-map.json"))
    parser.add_argument("--report", type=Path, default=Path(".tmp/ecost-sync/report.json"))
    args = parser.parse_args()

    try:
        build_dataset(args.input, args.existing, args.output, args.report)
    except Exception as exc:
        print(f"Network Map sync failed: {exc}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
